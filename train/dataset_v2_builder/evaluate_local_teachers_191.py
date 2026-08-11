from __future__ import annotations

import argparse
import json
from collections import Counter
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook

FAST_LABELS = [
    "objectif", "verrou", "methode", "parametre",
    "resultat", "limite", "contribution", "bruit"
]
VERROU_LABELS = ["verrou_evidence", "non_verrou"]

def sheet_rows(ws):
    headers = [c.value for c in ws[1]]
    out = []
    for vals in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, vals))
        if row.get("id"):
            out.append(row)
    return out

def load_gold(path: Path):
    wb = load_workbook(path, data_only=True)
    return {
        "fastjudge": sheet_rows(wb["FastJudge_96"]),
        "verrou": sheet_rows(wb["Verrou_95"]),
    }

def read_jsonl(path: Path) -> Dict[str, Dict[str, Any]]:
    out = {}
    if not path.exists():
        return out
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            try:
                row = json.loads(line)
                rid = str(row.get("id") or "")
                if rid:
                    out[rid] = row
            except Exception:
                pass
    return out

def metric_rows(gold_rows, predictions, pred_key, labels):
    pairs = []
    for row in gold_rows:
        truth = str(row.get("human_label") or "").strip()
        if truth not in labels:
            continue

        if predictions is None:
            pred = str(row.get(pred_key) or "").strip()
        else:
            pred_row = predictions.get(str(row["id"]))
            if not pred_row:
                continue
            pred = str(pred_row.get("label") or "").strip()

        if pred not in labels:
            continue
        pairs.append((truth, pred))

    if not pairs:
        return None

    correct = sum(t == p for t, p in pairs)
    per = {}
    f1s = []

    for lab in labels:
        tp = sum(t == lab and p == lab for t, p in pairs)
        fp = sum(t != lab and p == lab for t, p in pairs)
        fn = sum(t == lab and p != lab for t, p in pairs)

        precision = tp / (tp + fp) if tp + fp else 0.0
        recall = tp / (tp + fn) if tp + fn else 0.0
        f1 = 2 * precision * recall / (precision + recall) if precision + recall else 0.0
        support = sum(t == lab for t, _ in pairs)

        per[lab] = {
            "precision": round(precision, 4),
            "recall": round(recall, 4),
            "f1": round(f1, 4),
            "support": support,
        }
        f1s.append(f1)

    return {
        "n": len(pairs),
        "accuracy": round(correct / len(pairs), 4),
        "macro_f1": round(sum(f1s) / len(f1s), 4),
        "per_class": per,
    }

def print_result(task, model, res):
    print("\n" + "=" * 88)
    print(f"{task} — {model}")
    print("=" * 88)
    if not res:
        print("Aucun résultat exploitable.")
        return
    print("N         :", res["n"])
    print("Accuracy  :", f"{100*res['accuracy']:.2f}%")
    print("Macro-F1  :", f"{100*res['macro_f1']:.2f}%")
    print("Par classe :")
    for lab, m in res["per_class"].items():
        print(
            f"  {lab:18s} "
            f"P={m['precision']:.3f} "
            f"R={m['recall']:.3f} "
            f"F1={m['f1']:.3f} "
            f"N={m['support']}"
        )

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", default=r"C:\EnnoSmart")
    ap.add_argument(
        "--gold-xlsx",
        default=r"C:\EnnoSmart\train\data_v2\gold_191\GOLD_191_ANNOTATION_revu_GPT56.xlsx",
    )
    ap.add_argument(
        "--models",
        nargs="+",
        default=["qwen3:8b", "qwen2.5:7b-instruct"],
    )
    args = ap.parse_args()

    gold_path = Path(args.gold_xlsx)
    if not gold_path.exists():
        raise SystemExit(f"Fichier GOLD absent : {gold_path}")

    gold = load_gold(gold_path)
    out_dir = Path(args.root) / "train" / "data_v2" / "model_comparison_191"

    report = {"fastjudge": {}, "verrou": {}}

    # Baselines déjà dans le fichier GOLD
    for task, labels in [("fastjudge", FAST_LABELS), ("verrou", VERROU_LABELS)]:
        rows = gold[task]

        candidate_res = metric_rows(rows, None, "candidate_label", labels)
        qwen4_res = metric_rows(rows, None, "qwen3_label", labels)

        print_result(task.upper(), "candidate_label (ancien pré-label)", candidate_res)
        print_result(task.upper(), "qwen3:4b-instruct (déjà testé)", qwen4_res)

        report[task]["candidate_label"] = candidate_res
        report[task]["qwen3:4b-instruct"] = qwen4_res

        for model in args.models:
            safe = model.replace(":", "_").replace("/", "_")
            pred_file = out_dir / f"{task}_{safe}.jsonl"
            preds = read_jsonl(pred_file)
            res = metric_rows(rows, preds, "", labels)

            print_result(task.upper(), model, res)
            report[task][model] = res

    report_path = out_dir / "comparison_report.json"
    report_path.write_text(
        json.dumps(report, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print("\n" + "=" * 88)
    print("RAPPORT JSON :", report_path)
    print("=" * 88)

    # Résumé simple
    print("\nRESUME")
    for task in ("fastjudge", "verrou"):
        ranked = []
        for model, res in report[task].items():
            if res:
                ranked.append((res["macro_f1"], res["accuracy"], model))
        ranked.sort(reverse=True)
        print(f"\n{task.upper()}")
        for f1, acc, model in ranked:
            print(f"  {model:36s} Accuracy={100*acc:6.2f}%  Macro-F1={100*f1:6.2f}%")

if __name__ == "__main__":
    main()
