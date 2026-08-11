from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Dict, List, Any

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

FAST_LABELS = [
    "objectif","verrou","methode","parametre",
    "resultat","limite","contribution","bruit"
]
VERROU_LABELS = ["verrou_evidence","non_verrou"]

def read_csv(path: Path) -> List[Dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))

def read_jsonl(path: Path) -> List[Dict[str, Any]]:
    rows = []
    if not path.exists():
        return rows
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
                if isinstance(obj, dict):
                    rows.append(obj)
            except Exception:
                pass
    return rows

def index_by_id(rows):
    return {str(r.get("id")): r for r in rows if r.get("id")}

def build_rows(queue_rows, teacher_rows, task):
    q = index_by_id(queue_rows)
    out = []
    missing = []

    for t in teacher_rows:
        rid = str(t.get("id") or "")
        src = q.get(rid)
        if not src:
            missing.append(rid)
            continue

        out.append({
            "task": task,
            "id": rid,
            "source": src.get("source",""),
            "project_id": src.get("project_id",""),
            "title": src.get("title",""),
            "section_title": src.get("section_title",""),
            "candidate_label": src.get("candidate_label",""),
            "candidate_bucket": src.get("candidate_bucket",""),
            "qwen3_label": t.get("teacher_label",""),
            "qwen3_confidence": t.get("teacher_confidence",""),
            "qwen3_reason": t.get("teacher_reason",""),
            "text": src.get("text",""),
            "human_label": "",
            "human_comment": "",
        })
    return out, missing

def write_csv(path: Path, rows):
    fields = [
        "task","id","source","project_id","title","section_title",
        "candidate_label","candidate_bucket",
        "qwen3_label","qwen3_confidence","qwen3_reason",
        "text","human_label","human_comment"
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)

def add_sheet(ws, rows, allowed_labels):
    headers = [
        "task","id","source","project_id","title","section_title",
        "candidate_label","candidate_bucket",
        "qwen3_label","qwen3_confidence","qwen3_reason",
        "text","human_label","human_comment"
    ]
    ws.append(headers)

    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([row.get(h,"") for h in headers])

    # Large text columns
    widths = {
        1:16, 2:42, 3:10, 4:24, 5:32, 6:24,
        7:20, 8:20, 9:20, 10:16, 11:45, 12:90, 13:22, 14:40
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row in ws.iter_rows(min_row=2):
        row[10].alignment = Alignment(wrap_text=True, vertical="top")
        row[11].alignment = Alignment(wrap_text=True, vertical="top")
        row[13].alignment = Alignment(wrap_text=True, vertical="top")

    # Dropdown for human label
    formula = '"' + ",".join(allowed_labels) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    if len(rows) > 0:
        dv.add(f"M2:M{len(rows)+1}")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:N{len(rows)+1}"

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", default=r"C:\EnnoSmart")
    args = ap.parse_args()

    root = Path(args.root)
    qdir = root / "train" / "data_v2" / "teacher_queues"
    tdir = root / "train" / "data_v2" / "teacher_labels_local_stratified"
    outdir = root / "train" / "data_v2" / "gold_191"
    outdir.mkdir(parents=True, exist_ok=True)

    fast_queue = read_csv(qdir / "fastjudge_teacher_queue_v2.csv")
    verrou_queue = read_csv(qdir / "verrou_teacher_queue_v2.csv")

    fast_teacher = read_jsonl(tdir / "fastjudge_stratified_qwen3.jsonl")
    verrou_teacher = read_jsonl(tdir / "verrou_stratified_qwen3.jsonl")

    fast_rows, missing_fast = build_rows(fast_queue, fast_teacher, "fastjudge")
    verrou_rows, missing_verrou = build_rows(verrou_queue, verrou_teacher, "verrou_detector")

    fast_csv = outdir / "gold_fastjudge_96.csv"
    verrou_csv = outdir / "gold_verrou_95.csv"
    write_csv(fast_csv, fast_rows)
    write_csv(verrou_csv, verrou_rows)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "FastJudge_96"
    add_sheet(ws1, fast_rows, FAST_LABELS)

    ws2 = wb.create_sheet("Verrou_95")
    add_sheet(ws2, verrou_rows, VERROU_LABELS)

    ws3 = wb.create_sheet("Instructions")
    instructions = [
        ["Objectif", "Créer un petit GOLD humain indépendant pour comparer les anciens pré-labels et Qwen3."],
        ["Règle", "Ne choisis pas un label parce que candidate_label ou qwen3_label le propose. Lis le texte."],
        ["FastJudge", "Choisir UNE classe parmi : " + ", ".join(FAST_LABELS)],
        ["VerrouDetector", "Choisir verrou_evidence seulement s'il existe une vraie incertitude scientifique/technologique non maîtrisée."],
        ["non_verrou", "Objectif, méthode, résultat, bug, intégration, réglage, optimisation classique, contrainte, difficulté ordinaire ou limite connue."],
        ["À remplir", "Seulement human_label. human_comment est facultatif mais utile pour les cas difficiles."],
    ]
    for row in instructions:
        ws3.append(row)
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 120
    for row in ws3.iter_rows():
        row[1].alignment = Alignment(wrap_text=True, vertical="top")

    xlsx = outdir / "GOLD_191_ANNOTATION.xlsx"
    wb.save(xlsx)

    report = {
        "fastjudge_rows": len(fast_rows),
        "verrou_rows": len(verrou_rows),
        "total": len(fast_rows) + len(verrou_rows),
        "missing_fast_ids": missing_fast,
        "missing_verrou_ids": missing_verrou,
        "xlsx": str(xlsx),
        "fast_csv": str(fast_csv),
        "verrou_csv": str(verrou_csv),
    }
    (outdir / "gold_191_report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )
    print(json.dumps(report, ensure_ascii=False, indent=2))

if __name__ == "__main__":
    main()
