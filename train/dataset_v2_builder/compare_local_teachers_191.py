from __future__ import annotations

import argparse
import json
import time
from pathlib import Path
from typing import Any, Dict, List, Tuple

import requests
from openpyxl import load_workbook

OLLAMA_CHAT_URL = "http://127.0.0.1:11434/api/chat"
OLLAMA_TAGS_URL = "http://127.0.0.1:11434/api/tags"

FAST_LABELS = [
    "objectif", "verrou", "methode", "parametre",
    "resultat", "limite", "contribution", "bruit"
]
VERROU_LABELS = ["verrou_evidence", "non_verrou"]

FAST_SYSTEM = """Tu es annotateur expert de passages issus de projets R&D/CIR.
Classe chaque passage selon son sens principal dans UNE seule classe :

objectif = but, finalité ou résultat attendu
verrou = incertitude scientifique/technologique non résolue, obstacle de connaissance
methode = démarche, protocole, algorithme, expérimentation, procédé ou moyen utilisé
parametre = variable, réglage, configuration, seuil ou caractéristique mesurée
resultat = résultat obtenu, observation, mesure ou performance constatée
limite = faiblesse, restriction ou insuffisance connue, sans être forcément un verrou
contribution = apport, nouveauté, innovation ou solution proposée
bruit = administratif, contexte non pertinent, trop vague ou hors catégories

Règles strictes :
- une difficulté n'est pas automatiquement un verrou ;
- une limite connue n'est pas automatiquement un verrou ;
- pour "verrou", il faut une incertitude de connaissance réellement non maîtrisée ;
- n'utilise que les labels autorisés.
"""

VERROU_SYSTEM = """Tu es annotateur expert et strict pour les verrous R&D/CIR.

verrou_evidence = le passage contient une véritable incertitude scientifique ou technologique
non maîtrisée : phénomène, causalité, validité, représentativité, prédiction, généralisation,
transposabilité, comportement ou connaissance manquante.

non_verrou = objectif, méthode, résultat, bug, intégration, maintenance, réglage,
optimisation classique, contrainte client, difficulté d'implémentation, problème résolu,
ou limite connue sans incertitude scientifique réelle.

Une simple difficulté ou l'expression "surmonter une limite" ne suffit jamais à constituer
un verrou. N'utilise que verrou_evidence ou non_verrou.
"""

def sheet_rows(ws) -> List[Dict[str, Any]]:
    headers = [c.value for c in ws[1]]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        if row.get("id") and row.get("text"):
            rows.append(row)
    return rows

def load_gold(path: Path) -> Dict[str, List[Dict[str, Any]]]:
    wb = load_workbook(path, data_only=True)
    return {
        "fastjudge": sheet_rows(wb["FastJudge_96"]),
        "verrou": sheet_rows(wb["Verrou_95"]),
    }

def checkpoint_path(out_dir: Path, model: str, task: str) -> Path:
    safe = model.replace(":", "_").replace("/", "_")
    return out_dir / f"{task}_{safe}.jsonl"

def read_done(path: Path) -> Dict[str, Dict[str, Any]]:
    out = {}
    if not path.exists():
        return out
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            try:
                row = json.loads(line)
                rid = str(row.get("id") or "")
                if rid:
                    out[rid] = row
            except Exception:
                pass
    return out

def ollama_models() -> List[str]:
    r = requests.get(OLLAMA_TAGS_URL, timeout=10)
    r.raise_for_status()
    data = r.json()
    return [str(m.get("name") or "") for m in data.get("models", [])]

def schema_for(labels: List[str]) -> Dict[str, Any]:
    return {
        "type": "object",
        "properties": {
            "items": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "idx": {"type": "integer"},
                        "label": {"type": "string", "enum": labels},
                        "confidence": {"type": "number", "minimum": 0, "maximum": 1},
                        "reason": {"type": "string"},
                    },
                    "required": ["idx", "label", "confidence", "reason"],
                    "additionalProperties": False,
                },
            }
        },
        "required": ["items"],
        "additionalProperties": False,
    }

def ask_batch(
    model: str,
    task: str,
    batch: List[Dict[str, Any]],
    timeout: int,
    num_ctx: int,
) -> List[Dict[str, Any]]:
    labels = FAST_LABELS if task == "fastjudge" else VERROU_LABELS
    system = FAST_SYSTEM if task == "fastjudge" else VERROU_SYSTEM

    items = [
        {"idx": i + 1, "text": str(row.get("text") or "")[:1800]}
        for i, row in enumerate(batch)
    ]

    user = {
        "instruction": (
            "Classe TOUS les passages. Retourne exactement un objet pour chaque idx, "
            "sans en oublier et sans ajouter d'autre texte."
        ),
        "items": items,
    }

    payload = {
        "model": model,
        "stream": False,
        "keep_alive": "30m",
        "format": schema_for(labels),
        "options": {
            "temperature": 0,
            "num_ctx": num_ctx,
        },
        "messages": [
            {"role": "system", "content": system},
            {"role": "user", "content": json.dumps(user, ensure_ascii=False)},
        ],
    }

    r = requests.post(OLLAMA_CHAT_URL, json=payload, timeout=timeout)
    r.raise_for_status()

    content = str(r.json().get("message", {}).get("content") or "").strip()
    data = json.loads(content)
    preds = data.get("items", [])
    if not isinstance(preds, list):
        raise ValueError("Réponse structurée invalide : items n'est pas une liste.")
    return preds

def ask_one(model: str, task: str, row: Dict[str, Any], timeout: int, num_ctx: int):
    preds = ask_batch(model, task, [row], timeout, num_ctx)
    if not preds:
        return None
    return preds[0]

def annotate_task(
    model: str,
    task: str,
    rows: List[Dict[str, Any]],
    out_path: Path,
    batch_size: int,
    timeout: int,
    num_ctx: int,
):
    allowed = set(FAST_LABELS if task == "fastjudge" else VERROU_LABELS)
    done = read_done(out_path)
    remaining = [r for r in rows if str(r["id"]) not in done]

    print(f"\n[{model}] task={task} total={len(rows)} deja={len(done)} restant={len(remaining)}")

    with out_path.open("a", encoding="utf-8") as f:
        for start in range(0, len(remaining), batch_size):
            batch = remaining[start:start + batch_size]
            batch_no = start // batch_size + 1

            preds = []
            try:
                preds = ask_batch(model, task, batch, timeout, num_ctx)
            except Exception as exc:
                print(f"[batch {batch_no}] fallback individuel : {exc!r}")

            by_idx: Dict[int, Dict[str, Any]] = {}
            for p in preds:
                try:
                    idx = int(p.get("idx"))
                except Exception:
                    continue
                if 1 <= idx <= len(batch):
                    by_idx[idx] = p

            written = 0
            for local_idx, row in enumerate(batch, start=1):
                p = by_idx.get(local_idx)

                if p is None:
                    try:
                        p = ask_one(model, task, row, timeout, num_ctx)
                    except Exception as exc:
                        print(f"  [ECHEC] {row['id']} : {exc!r}")
                        continue

                label = str(p.get("label") or "").strip()
                if label not in allowed:
                    print(f"  [LABEL INVALIDE] {row['id']} -> {label!r}")
                    continue

                try:
                    confidence = float(p.get("confidence") or 0)
                except Exception:
                    confidence = 0.0

                rec = {
                    "id": str(row["id"]),
                    "task": task,
                    "model": model,
                    "label": label,
                    "confidence": max(0.0, min(1.0, confidence)),
                    "reason": str(p.get("reason") or "")[:500],
                }
                f.write(json.dumps(rec, ensure_ascii=False) + "\n")
                f.flush()
                written += 1

            print(
                f"[OK] {task} batch={batch_no} "
                f"requested={len(batch)} written={written}"
            )

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", default=r"C:\EnnoSmart")
    ap.add_argument(
        "--gold-xlsx",
        default=r"C:\EnnoSmart\train\data_v2\gold_191\GOLD_191_ANNOTATION_revu_GPT56.xlsx",
    )
    ap.add_argument(
        "--models",
        nargs="+",
        default=["qwen3:8b", "qwen2.5:7b-instruct"],
    )
    ap.add_argument("--batch-size", type=int, default=4)
    ap.add_argument("--timeout", type=int, default=600)
    ap.add_argument("--num-ctx", type=int, default=8192)
    args = ap.parse_args()

    gold_path = Path(args.gold_xlsx)
    if not gold_path.exists():
        raise SystemExit(
            f"Fichier GOLD absent : {gold_path}\n"
            "Passe son chemin avec --gold-xlsx."
        )

    installed = ollama_models()
    print("[Ollama] modèles installés :", installed)

    for model in args.models:
        if model not in installed:
            print(f"[ATTENTION] {model} n'apparaît pas exactement dans ollama list.")
            print("Le script va quand même essayer ce nom.")

    gold = load_gold(gold_path)
    out_dir = Path(args.root) / "train" / "data_v2" / "model_comparison_191"
    out_dir.mkdir(parents=True, exist_ok=True)

    print("[GOLD] FastJudge :", len(gold["fastjudge"]))
    print("[GOLD] VerrouDetector :", len(gold["verrou"]))

    for model in args.models:
        annotate_task(
            model,
            "fastjudge",
            gold["fastjudge"],
            checkpoint_path(out_dir, model, "fastjudge"),
            args.batch_size,
            args.timeout,
            args.num_ctx,
        )
        annotate_task(
            model,
            "verrou",
            gold["verrou"],
            checkpoint_path(out_dir, model, "verrou"),
            args.batch_size,
            args.timeout,
            args.num_ctx,
        )

    print("\n[TERMINE] Predictions sauvegardées dans :", out_dir)

if __name__ == "__main__":
    main()
