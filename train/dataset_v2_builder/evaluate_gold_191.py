from __future__ import annotations

import argparse
from collections import Counter, defaultdict
from pathlib import Path
from typing import List, Dict, Any

from openpyxl import load_workbook

FAST_LABELS = [
    "objectif","verrou","methode","parametre",
    "resultat","limite","contribution","bruit"
]
VERROU_LABELS = ["verrou_evidence","non_verrou"]

def sheet_rows(ws):
    headers = [c.value for c in ws[1]]
    out = []
    for vals in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, vals))
        if row.get("id"):
            out.append(row)
    return out

def metrics(rows: List[Dict[str, Any]], pred_key: str, labels: List[str]):
    valid = [r for r in rows if str(r.get("human_label") or "") in labels]
    if not valid:
        return None

    accuracy = sum(str(r.get(pred_key)) == str(r.get("human_label")) for r in valid) / len(valid)

    per = {}
    f1s = []
    for lab in labels:
        tp = sum(r.get(pred_key) == lab and r.get("human_label") == lab for r in valid)
        fp = sum(r.get(pred_key) == lab and r.get("human_label") != lab for r in valid)
        fn = sum(r.get(pred_key) != lab and r.get("human_label") == lab for r in valid)

        precision = tp / (tp + fp) if tp + fp else 0.0
        recall = tp / (tp + fn) if tp + fn else 0.0
        f1 = 2*precision*recall/(precision+recall) if precision+recall else 0.0
        f1s.append(f1)
        per[lab] = {
            "precision": round(precision,4),
            "recall": round(recall,4),
            "f1": round(f1,4),
            "support": sum(r.get("human_label") == lab for r in valid)
        }

    return {
        "n": len(valid),
        "accuracy": round(accuracy,4),
        "macro_f1": round(sum(f1s)/len(f1s),4),
        "per_class": per,
    }

def print_block(title, result):
    print("\n" + "="*80)
    print(title)
    print("="*80)
    if result is None:
        print("Aucun human_label valide rempli.")
        return
    print("N =", result["n"])
    print("Accuracy =", result["accuracy"])
    print("Macro-F1 =", result["macro_f1"])
    print("Par classe :")
    for lab, m in result["per_class"].items():
        print(
            f"  {lab:18s} "
            f"P={m['precision']:.4f} "
            f"R={m['recall']:.4f} "
            f"F1={m['f1']:.4f} "
            f"N={m['support']}"
        )

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", default=r"C:\EnnoSmart")
    args = ap.parse_args()

    xlsx = Path(args.root) / "train" / "data_v2" / "gold_191" / "GOLD_191_ANNOTATION.xlsx"
    if not xlsx.exists():
        raise SystemExit(f"Fichier absent : {xlsx}")

    wb = load_workbook(xlsx, data_only=True)
    fast = sheet_rows(wb["FastJudge_96"])
    verrou = sheet_rows(wb["Verrou_95"])

    print_block(
        "FASTJUDGE — ancien candidat vs GOLD humain",
        metrics(fast, "candidate_label", FAST_LABELS)
    )
    print_block(
        "FASTJUDGE — Qwen3 vs GOLD humain",
        metrics(fast, "qwen3_label", FAST_LABELS)
    )
    print_block(
        "VERROU — ancien candidat vs GOLD humain",
        metrics(verrou, "candidate_label", VERROU_LABELS)
    )
    print_block(
        "VERROU — Qwen3 vs GOLD humain",
        metrics(verrou, "qwen3_label", VERROU_LABELS)
    )

if __name__ == "__main__":
    main()
