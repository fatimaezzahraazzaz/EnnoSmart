"""
modules/extraction/structured/excel_struct.py

Extraction structuré de fichiers Excel pour dossiers R&D / CIR.

Double sortie (conformément à l'architecture EnnoSmart) :

  1. text_chunks  RAG (tous les agents)
     Chaque feuille a un chunk texte avec tableaux en Markdown,
     labels, commentaires de cellules, et valeurs numériques clés.

  2. structured_data Enno Valo (mapping Excel/Cerfa)
     Données brutes structurées : cellules nommées, tableaux détectés,
     totaux, en-têtes détectés, pour remplissage automatique de templates.

Stratégie d'extraction :
  - openpyxl (data_only=True) : valeurs calculées + métadonnées cellules
  - Détection automatique des tableaux (zones rectangulaires denses)
  - Extraction des commentaires/notes de cellules (annotations consultant)
  - Détection des cellules nommées (Named Ranges) → données clés CIR
  - Reconnaissance de patterns R&D : budgets, personnel, dépenses éligibles
  - Reconstruction texte fidèle pour le RAG
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional, Union

from modules.extraction.formula.formula import extract_formulas
try:
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.worksheet.worksheet import Worksheet
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

logger = logging.getLogger(__name__)


# Nombre minimum de cellules non-vides pour considÃ©rer une ligne comme "donnÃ©e"
MIN_CELLS_FOR_DATA_ROW = 2

# Nombre minimum de lignes pour dÃ©tecter un tableau
MIN_ROWS_FOR_TABLE = 2

# Seuil de densité pour la detection de tableau (% cellules non-vides)
TABLE_DENSITY_THRESHOLD = 0.40

# Valeur max de caracteres par cellule pour le texte RAG
MAX_CELL_TEXT_LENGTH = 500

# Patterns de libellés R&D typiques dans les Excel CIR
RD_LABEL_PATTERNS: dict[str, list[str]] = {
    "PERSONNEL_RD": [
        "personnel r&d", "chercheurs", "ingénieurs r&d",
        "doctorants", "personnel de recherche", "etp r&d",
        "equivalent temps plein", "salaires r&d",
    ],
    "DEPENSES_RD": [
        "dÃ©penses r&d", "dÃ©penses Ã©ligibles", "base cir",
        "montant cir", "crÃ©dit impÃ´t recherche",
        "sous-traitance r&d", "frais de brevets",
        "dotations amortissements", "frais de fonctionnement",
    ],
    "BUDGET_PROJET": [
        "budget total", "coÃ»t total", "financement",
        "budget prÃ©visionnel", "dÃ©penses totales",
        "recettes", "subvention", "apport propre",
    ],
    "RESULTATS_RD": [
        "rÃ©sultats", "indicateurs", "kpi", "livrables",
        "jalons", "avancement", "taux rÃ©alisation",
        "publications", "brevets dÃ©posÃ©s",
    ],
    "DATES_PROJET": [
        "date dÃ©but", "date fin", "durÃ©e", "pÃ©riode",
        "exercice", "annÃ©e fiscale", "trimestre",
    ],
}

# Patterns sections R&D pour le RAG (identiques aux autres extracteurs)
RD_SECTION_PATTERNS: list[str] = [
    r"(objectif[s]?\s+(?:du\s+)?projet)",
    r"(Ã©tat\s+de\s+l['\s]art)",
    r"(verrous?\s+technologique[s]?)",
    r"(travaux\s+r(?:echerche)?(?:\s*&\s*|\s+et\s+)d(?:Ã©veloppement)?)",
    r"(rÃ©sultats?\s+(?:obtenus?|attendus?))",
    r"(dÃ©penses?\s+(?:de\s+)?recherche)",
    r"(personnel\s+(?:de\s+)?recherche)",
    r"(budget\s+(?:r&d|recherche|projet))",
]

_RD_SECTION_RE = re.compile(
    "|".join(RD_SECTION_PATTERNS),
    flags=re.IGNORECASE | re.UNICODE,
)

# Organismes dÃ©tectables dans les noms de feuilles / cellules

# - Dataclasses -------------------------------

@dataclass
class CellValue:
    """Représentation d'une cellule avec toutes ses métadonnées."""
    row: int
    col: int
    col_letter: str
    address: str                        # Ex: "B5"
    value: Any                          # Valeur brute
    value_str: str                      # Valeur convertie en str propre
    data_type: str                      # "text" | "number" | "date" | "formula" | "empty"
    comment: Optional[str]             # Note/commentaire de cellule
    is_bold: bool                       # Souvent = en-tête ou total
    is_merged: bool
    number_format: Optional[str]        # Format Excel (â‚¬, %, dateâ€¦)


@dataclass
class DetectedTable:
    """Un tableau rectangulaire détecté dans une feuille."""
    sheet_name: str
    start_row: int
    end_row: int
    start_col: int
    end_col: int
    address_range: str                  # Ex: "B3:F15"
    headers: list[str]                  # Première ligne du tableau
    rows_data: list[list[str]]          # Données (hors en-tête)
    row_count: int
    col_count: int
    markdown: str                       # Rendu Markdown pour le RAG
    rd_category: Optional[str]         # PERSONNEL_RD, DEPENSES_RDâ€¦
    total_row: Optional[list[str]]     # Dernière ligne si c'est un total


@dataclass
class NamedRange:
    """Plage nommée Excel (données clés pour Valo)."""
    name: str
    sheet_name: Optional[str]
    address: str
    value: Any
    value_str: str
    rd_relevance: Optional[str]         # Catégorie R&D si détectée


@dataclass
class SheetResult:
    """Résultat d'extraction pour une feuille Excel."""
    sheet_name: str
    sheet_index: int
    is_hidden: bool

    # - Pour le RAG ---------------------------â”€
    text_chunk: str                     # Chunk complet de la feuille
    cell_comments: list[str]           # Commentaires de cellules
    detected_sections: list[str]       # Sections R&D détectées

    # - Pour Valo (structuré) ----------------------â”€
    tables: list[DetectedTable]
    raw_cells: dict[str, CellValue]    # Adresse → CellValue (cellules non-vides)
    key_cells: dict[str, CellValue]    # Cellules identifiées comme données CIR clés

    # - Statistiques ---------------------------
    total_rows: int
    total_cols: int
    non_empty_cells: int
    has_numbers: bool
    has_dates: bool
    rd_categories_found: list[str]


@dataclass
class ExcelStructResult:
    """
    RÃ©sultat complet d'extraction structurÃ©e Excel.
    Compatible avec ExtractionResult (base.py).
    """
    file_name: str
    source_path: str
    file_type: str = "excel"

    # - Sortie 1 : RAG (tous agents) -------------------â”€
    text_chunks: list[str] = field(default_factory=list)
    # Un chunk par feuille non-vide

    # - Sortie 2 : StructurÃ© (Enno Valo) -----------------
    structured_data: dict = field(default_factory=dict)
    # {
    #   "sheets": [...],
    #   "named_ranges": [...],
    #   "all_tables": [...],
    #   "key_values": {...}    â† valeurs CIR clÃ©s extraites
    # }

    # - DÃ©tail par feuille ------------------------â”€
    sheets: list[SheetResult] = field(default_factory=list)

    # - Plages nommÃ©es --------------------------â”€
    named_ranges: list[NamedRange] = field(default_factory=list)

    # - MÃ©tadonnÃ©es document -----------------------â”€
    sheet_names: list[str] = field(default_factory=list)
    organisme_detected: Optional[str] = None
    detected_rd_sections: list[str] = field(default_factory=list)

    # - TraÃ§abilitÃ© ----------------------------
    tags: list[str] = field(default_factory=list)
    confidence_score: float = 1.0
    extraction_errors: list[str] = field(default_factory=list)


# - Utilitaires -------------------------------

def _cell_value_to_str(value: Any, number_format: Optional[str] = None) -> str:
    """
    Convertit une valeur de cellule Excel en chaîne lisible.
    Gère les types : datetime, float, int, bool, str, None.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Oui" if value else "Non"
    if hasattr(value, "strftime"):
        # datetime ou date
        try:
            return value.strftime("%d/%m/%Y")
        except Exception:
            return str(value)
    if isinstance(value, float):
        # Éviter les flottants inutiles (ex: 30000.0 → "30 000")
        if value == int(value):
            return f"{int(value):,}".replace(",", " ")
        return f"{value:,.2f}".replace(",", " ")
    if isinstance(value, int):
        return f"{value:,}".replace(",", " ")
    return str(value).strip()[:MAX_CELL_TEXT_LENGTH]


def _detect_data_type(value: Any) -> str:
    """Classifie le type de donnée d'une cellule."""
    if value is None:
        return "empty"
    if isinstance(value, bool):
        return "boolean"
    if hasattr(value, "strftime"):
        return "date"
    if isinstance(value, (int, float)):
        return "number"
    if isinstance(value, str) and value.startswith("="):
        return "formula"
    return "text"


def _detect_rd_category(label: str) -> Optional[str]:
    """
    Détecte la catégorie R&D d'un libellé de cellule/en-tête.
    Retourne la clé de RD_LABEL_PATTERNS ou None.
    """
    label_lower = label.lower().strip()
    for category, keywords in RD_LABEL_PATTERNS.items():
        if any(kw in label_lower for kw in keywords):
            return category
    return None


def _detect_rd_sections(text: str) -> list[str]:
    matches = _RD_SECTION_RE.findall(text)
    sections = []
    for match in matches:
        if isinstance(match, tuple):
            sections.extend(s.strip() for s in match if s.strip())
        elif match.strip():
            sections.append(match.strip())
    seen: set[str] = set()
    result: list[str] = []
    for s in sections:
        if s.lower() not in seen:
            seen.add(s.lower())
            result.append(s)
    return result


def _table_to_markdown(headers: list[str], rows: list[list[str]]) -> str:
    """Convertit un tableau en Markdown à identique aux autres extracteurs."""
    if not headers:
        return ""
    all_rows = [headers] + rows
    col_count = max(len(r) for r in all_rows)
    col_widths = [
        max((len(r[i]) if i < len(r) else 0) for r in all_rows)
        for i in range(col_count)
    ]
    col_widths = [max(w, 3) for w in col_widths]

    def _fmt(row: list[str]) -> str:
        padded = [
            (row[i].ljust(col_widths[i]) if i < len(row) else " " * col_widths[i])
            for i in range(col_count)
        ]
        return "| " + " | ".join(padded) + " |"

    sep = "| " + " | ".join("-" * w for w in col_widths) + " |"
    return "\n".join([_fmt(headers), sep] + [_fmt(r) for r in rows])


# - Extraction cellules ---------------------------

def _extract_cells(ws: "Worksheet") -> dict[str, CellValue]:
    """
    Extrait toutes les cellules non-vides d'une feuille openpyxl.
    Inclut : valeur, type, commentaire, formatage (bold, merged, format).
    """
    cells: dict[str, CellValue] = {}

    # Cellules fusionnées : on récupère les plages pour le flag is_merged
    merged_ranges: set[str] = set()
    for merged_range in ws.merged_cells.ranges:
        for row in ws.iter_rows(
            min_row=merged_range.min_row, max_row=merged_range.max_row,
            min_col=merged_range.min_col, max_col=merged_range.max_col,
        ):
            for cell in row:
                merged_ranges.add(cell.coordinate)

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue

            col_letter = get_column_letter(cell.column)
            address = f"{col_letter}{cell.row}"
            value_str = _cell_value_to_str(cell.value, cell.number_format)

            if not value_str.strip():
                continue

            # Commentaire de cellule
            comment_text: Optional[str] = None
            if cell.comment and cell.comment.text:
                comment_text = cell.comment.text.strip()

            # Bold : souvent indicateur d'en-tÃªte ou de total
            is_bold = False
            try:
                is_bold = bool(cell.font and cell.font.bold)
            except Exception:
                pass

            cells[address] = CellValue(
                row=cell.row,
                col=cell.column,
                col_letter=col_letter,
                address=address,
                value=cell.value,
                value_str=value_str,
                data_type=_detect_data_type(cell.value),
                comment=comment_text,
                is_bold=is_bold,
                is_merged=address in merged_ranges,
                number_format=cell.number_format,
            )

    return cells


# - Détection de tableaux --------------------------

def _detect_tables(
    ws: "Worksheet",
    cells: dict[str, CellValue],
) -> list[DetectedTable]:
    """
    Détecte les zones rectangulaires denses de données dans une feuille.

    Algorithme :
      1. Trouve les lignes avec au moins MIN_CELLS_FOR_DATA_ROW cellules
      2. Regroupe les lignes consécutives en blocs
      3. Pour chaque bloc de taille suffisante → tableau détecté
      4. La première ligne du bloc = en-têtes (si bold ou texte)
      5. Détecte si la dernière ligne est un total (bold, "total", "somme"…)
    """
    sheet_name = ws.title

    # Analyser ligne par ligne
    row_data: dict[int, list[CellValue]] = {}
    for address, cell in cells.items():
        if cell.row not in row_data:
            row_data[cell.row] = []
        row_data[cell.row].append(cell)

    # Trier par ligne
    sorted_rows = sorted(row_data.items())

    # Trouver les blocs de lignes consécutives avec données
    blocks: list[list[int]] = []
    current_block: list[int] = []

    for row_num, row_cells in sorted_rows:
        if len(row_cells) >= MIN_CELLS_FOR_DATA_ROW:
            if current_block and row_num - current_block[-1] > 2:
                # Saut > 2 lignes → nouveau bloc
                if len(current_block) >= MIN_ROWS_FOR_TABLE:
                    blocks.append(current_block)
                current_block = []
            current_block.append(row_num)
        elif current_block:
            if len(current_block) >= MIN_ROWS_FOR_TABLE:
                blocks.append(current_block)
            current_block = []

    if current_block and len(current_block) >= MIN_ROWS_FOR_TABLE:
        blocks.append(current_block)

    tables: list[DetectedTable] = []

    for block_rows in blocks:
        # Trouver les colonnes impliquÃ©es
        block_cells = [
            cell for address, cell in cells.items()
            if cell.row in block_rows
        ]
        if not block_cells:
            continue

        min_col = min(c.col for c in block_cells)
        max_col = max(c.col for c in block_cells)
        min_row = min(block_rows)
        max_row = max(block_rows)

        # Construire la grille
        grid: dict[tuple[int, int], str] = {}
        for cell in block_cells:
            grid[(cell.row, cell.col)] = cell.value_str

        # Extraire les lignes complÃ¨tes
        all_rows: list[list[str]] = []
        for r in range(min_row, max_row + 1):
            row = [
                grid.get((r, c), "")
                for c in range(min_col, max_col + 1)
            ]
            all_rows.append(row)

        if not all_rows:
            continue

        # En-têtes = première ligne
        headers = all_rows[0]
        data_rows = all_rows[1:]

        # Vérifier densité minimale
        total_cells = len(all_rows) * (max_col - min_col + 1)
        filled_cells = sum(1 for r in all_rows for v in r if v.strip())
        density = filled_cells / total_cells if total_cells > 0 else 0

        if density < TABLE_DENSITY_THRESHOLD:
            continue

        # Ligne total : dernière ligne avec "total" ou "somme" ou en bold
        total_row: Optional[list[str]] = None
        if data_rows:
            last_row = data_rows[-1]
            last_row_str = " ".join(last_row).lower()
            if re.search(r"\b(total|somme|sum|sous-total|montant)\b", last_row_str):
                total_row = last_row
                data_rows = data_rows[:-1]

        # Catégorie R&D
        headers_text = " ".join(headers)
        rd_category = _detect_rd_category(headers_text)
        if not rd_category:
            # Chercher dans les libellés de la première colonne
            first_col_labels = " ".join(r[0] for r in data_rows if r)
            rd_category = _detect_rd_category(first_col_labels)

        start_col_letter = get_column_letter(min_col)
        end_col_letter = get_column_letter(max_col)
        address_range = f"{start_col_letter}{min_row}:{end_col_letter}{max_row}"

        tables.append(DetectedTable(
            sheet_name=sheet_name,
            start_row=min_row,
            end_row=max_row,
            start_col=min_col,
            end_col=max_col,
            address_range=address_range,
            headers=headers,
            rows_data=data_rows,
            row_count=len(data_rows),
            col_count=max_col - min_col + 1,
            markdown=_table_to_markdown(headers, data_rows),
            rd_category=rd_category,
            total_row=total_row,
        ))

    return tables


# - Identification des cellules clés CIR ------------------

def _identify_key_cells(cells: dict[str, CellValue]) -> dict[str, CellValue]:
    """
    Identifie les cellules contenant des données clés CIR.

    """
    key_cells: dict[str, CellValue] = {}

    # Index par position (row, col)
    by_pos: dict[tuple[int, int], CellValue] = {
        (c.row, c.col): c for c in cells.values()
    }

    for cell in cells.values():
        if cell.data_type != "text":
            continue

        rd_cat = _detect_rd_category(cell.value_str)
        if not rd_cat:
            continue

        # Chercher une valeur numérique dans les 3 colonnes suivantes
        for offset in range(1, 4):
            neighbor = by_pos.get((cell.row, cell.col + offset))
            if neighbor and neighbor.data_type == "number":
                key_cells[f"{cell.address}→{neighbor.address}"] = neighbor
                break

        # Ou dans la ligne suivante même colonne
        neighbor_below = by_pos.get((cell.row + 1, cell.col))
        if neighbor_below and neighbor_below.data_type == "number":
            key_cells[f"{cell.address}↓{neighbor_below.address}"] = neighbor_below

    return key_cells


# - Extraction plages nommées ------------------------
def _iter_defined_names(wb: "openpyxl.Workbook"):
    """
    Compatibilité openpyxl ancien / nouveau.
    Évite l'erreur :
    'DefinedNameDict' object has no attribute 'definedName'
    """
    defined_names = getattr(wb, "defined_names", None)

    if not defined_names:
        return []

    if hasattr(defined_names, "definedName"):
        return defined_names.definedName or []

    try:
        return list(defined_names.values())
    except Exception:
        return []
def _extract_named_ranges(wb: "openpyxl.Workbook") -> list[NamedRange]:
    """
    Extrait les plages nommées Excel de manière compatible avec
    plusieurs versions de openpyxl.
    """
    named_ranges: list[NamedRange] = []

    for named_range in _iter_defined_names(wb):
        try:
            name = (
                getattr(named_range, "name", None)
                or getattr(named_range, "localSheetId", None)
                or ""
            )

            if not name:
                continue

            destinations = []
            try:
                destinations = list(named_range.destinations)
            except Exception:
                destinations = []

            if not destinations:
                continue

            for sheet_title, coord in destinations:
                try:
                    ws = wb[sheet_title]
                except Exception:
                    continue

                value = None
                value_str = ""

                try:
                    clean_coord = str(coord).replace("$", "")

                    if ":" in clean_coord:
                        cells = ws[clean_coord]
                        values = []
                        for row in cells:
                            for cell in row:
                                if cell.value is not None:
                                    values.append(_cell_value_to_str(cell.value))
                        value = values
                        value_str = " | ".join(values)
                    else:
                        cell = ws[clean_coord]
                        value = cell.value
                        value_str = _cell_value_to_str(value)

                except Exception:
                    value = None
                    value_str = ""

                rd_relevance = _detect_rd_category(str(name))

                named_ranges.append(
                    NamedRange(
                        name=str(name),
                        sheet_name=sheet_title,
                        address=str(coord),
                        value=value,
                        value_str=value_str,
                        rd_relevance=rd_relevance,
                    )
                )

        except Exception as exc:
            logger.debug("Plage nommée ignorée : %s", exc)

    return named_ranges


# - Construction des chunks RAG -----------------------

def _build_sheet_chunk(
    sheet_name: str,
    tables: list[DetectedTable],
    cells: dict[str, CellValue],
    comments: list[str],
) -> str:
    """
    Construit le chunk RAG d'une feuille Excel.
    """
    parts: list[str] = [f"[FEUILLE : {sheet_name}]"]

    # - Tableaux -----------------------------
    for i, table in enumerate(tables, start=1):
        cat_str = f" | {table.rd_category}" if table.rd_category else ""
        header_line = f"[TABLEAU {i}{cat_str} | {table.address_range}]"
        parts.append(f"{header_line}\n{table.markdown}")

        # Ligne total si présente
        if table.total_row:
            total_str = " | ".join(v for v in table.total_row if v)
            parts.append(f"[TOTAL] {total_str}")

    # - Cellules isolées hors tableaux ------------------
    # Récupérer les cellules non couvertes par des tableaux
    table_rows: set[int] = set()
    for t in tables:
        table_rows.update(range(t.start_row, t.end_row + 1))

    isolated: list[str] = []
    by_row: dict[int, list[CellValue]] = {}
    for cell in cells.values():
        if cell.row not in table_rows:
            by_row.setdefault(cell.row, []).append(cell)

    for row_num in sorted(by_row.keys()):
        row_cells = sorted(by_row[row_num], key=lambda c: c.col)
        # Paires libellé : valeur
        texts = [c.value_str for c in row_cells if c.value_str.strip()]
        if texts:
            isolated.append(" | ".join(texts))

    if isolated:
        parts.append("[DONNÉES]\n" + "\n".join(isolated))

    # - Commentaires ---------------------------
    if comments:
        parts.append("[COMMENTAIRES CELLULES]\n" + "\n".join(f"• {c}" for c in comments))

    return "\n\n".join(parts)


# - Pipeline principal openpyxl -----------------------

def _extract_with_openpyxl(path: Path) -> ExcelStructResult:
    """Pipeline d'extraction complet via openpyxl."""
    result = ExcelStructResult(
        file_name=path.name,
        source_path=str(path.resolve()),
    )

    try:
        # data_only=True â†’ valeurs calculÃ©es (pas les formules)
        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=False)
    except Exception as exc:
        result.extraction_errors.append(f"Impossible d'ouvrir le fichier : {exc}")
        return result

    result.sheet_names = wb.sheetnames
    all_text_parts: list[str] = []
    all_rd_sections: set[str] = set()
    all_tables: list[DetectedTable] = []

    # - Plages nommÃ©es --------------------------
    try:
        result.named_ranges = _extract_named_ranges(wb)
    except Exception as exc:
        logger.warning("Erreur extraction plages nommÃ©es : %s", exc)

    # - Feuille par feuille -----------------------â”€
    for sheet_index, sheet_name in enumerate(wb.sheetnames):
        try:
            ws = wb[sheet_name]

            # VisibilitÃ©
            is_hidden = ws.sheet_state != "visible"

            # Extraction cellules
            cells = _extract_cells(ws)
            if not cells:
                logger.debug("Feuille '%s' vide â€” ignorÃ©e", sheet_name)
                continue

            # Commentaires
            comments = [
                f"{addr}: {cell.comment}"
                for addr, cell in cells.items()
                if cell.comment
            ]

            # DÃ©tection tableaux
            tables = _detect_tables(ws, cells)
            all_tables.extend(tables)

            # Cellules clÃ©s CIR
            key_cells = _identify_key_cells(cells)

            # Chunk RAG
            chunk = _build_sheet_chunk(sheet_name, tables, cells, comments)
            result.text_chunks.append(chunk)
            
            # Extraction des formules du chunk Excel
            formulas = extract_formulas(text=chunk)
            for formula in formulas:
                result.text_chunks.append(formula.rag_chunk)
            
            all_text_parts.append(chunk)

            # Sections R&D
            sections = _detect_rd_sections(chunk)
            all_rd_sections.update(sections)

            # CatÃ©gories R&D dÃ©tectÃ©es
            rd_categories = list({
                t.rd_category for t in tables if t.rd_category
            })

            # Statistiques
            number_cells = [c for c in cells.values() if c.data_type == "number"]
            date_cells = [c for c in cells.values() if c.data_type == "date"]

            sheet_result = SheetResult(
                sheet_name=sheet_name,
                sheet_index=sheet_index,
                is_hidden=is_hidden,
                text_chunk=chunk,
                cell_comments=comments,
                detected_sections=sections,
                tables=tables,
                raw_cells=cells,
                key_cells=key_cells,
                total_rows=ws.max_row or 0,
                total_cols=ws.max_column or 0,
                non_empty_cells=len(cells),
                has_numbers=bool(number_cells),
                has_dates=bool(date_cells),
                rd_categories_found=rd_categories,
            )

            result.sheets.append(sheet_result)

            logger.debug(
                "Feuille '%s' â€” %d cellules | %d tableaux | catÃ©gories=%s",
                sheet_name, len(cells), len(tables), rd_categories,
            )

        except Exception as exc:
            msg = f"Erreur feuille '{sheet_name}' : {exc}"
            logger.warning(msg)
            result.extraction_errors.append(msg)
            result.text_chunks.append(f"[FEUILLE : {sheet_name}]\n[ERREUR EXTRACTION]")

    # - Sections R&D globales ----------------------â”€
    result.detected_rd_sections = sorted(all_rd_sections)

    # - Organisme (reserve pour NER) -----------------------â”€
    full_text = " ".join(all_text_parts)
    result.organisme_detected = None

    # - structured_data (pour Enno Valo) -----------------
    result.structured_data = {
        "sheets": [
            {
                "name": s.sheet_name,
                "is_hidden": s.is_hidden,
                "tables": [
                    {
                        "range":       t.address_range,
                        "headers":     t.headers,
                        "rows":        t.rows_data,
                        "total_row":   t.total_row,
                        "rd_category": t.rd_category,
                        "row_count":   t.row_count,
                    }
                    for t in s.tables
                ],
                "key_cells": {
                    ref: {"address": cv.address, "value": cv.value_str}
                    for ref, cv in s.key_cells.items()
                },
                "rd_categories": s.rd_categories_found,
            }
            for s in result.sheets
        ],
        "named_ranges": [
            {
                "name":         nr.name,
                "sheet":        nr.sheet_name,
                "address":      nr.address,
                "value":        nr.value_str,
                "rd_relevance": nr.rd_relevance,
            }
            for nr in result.named_ranges
        ],
        "all_tables_count": len(all_tables),
        "organisme":        result.organisme_detected,
    }

    wb.close()
    return result


# - Tags & Confiance -----------------------------

def _build_tags(result: ExcelStructResult) -> list[str]:
    tags: list[str] = ["EXCEL"]

    all_categories: set[str] = set()
    has_tables = False
    has_comments = False

    for sheet in result.sheets:
        if sheet.tables:
            has_tables = True
        if sheet.cell_comments:
            has_comments = True
        all_categories.update(sheet.rd_categories_found)

    if has_tables:
        tags.append("HAS_TABLES")
    if has_comments:
        tags.append("HAS_CELL_COMMENTS")
    if result.named_ranges:
        tags.append("HAS_NAMED_RANGES")
    if result.detected_rd_sections:
        tags.append("CIR_SECTIONS")

    for cat in sorted(all_categories):
        tags.append(f"RD:{cat}")

    if result.extraction_errors:
        tags.append("PARTIAL_EXTRACTION")

    return tags


def _compute_confidence(result: ExcelStructResult) -> float:
    """
    Score de confiance basÃ© sur la richesse des donnÃ©es extraites.
    PÃ©nalise les feuilles vides, les erreurs, l'absence de tableaux.
    """
    if not result.sheets:
        return 0.10

    error_penalty = min(len(result.extraction_errors) * 0.10, 0.30)
    no_table_penalty = 0.10 if not any(s.tables for s in result.sheets) else 0.0
    score = 1.0 - error_penalty - no_table_penalty

    return max(round(score, 2), 0.10)


# - Point d'entrÃ©e principal -------------------------

def extract_excel(file_path: str | Path) -> ExcelStructResult:
    """
    Extrait le contenu d'un fichier Excel pour le RAG et Enno Valo.

    Double sortie :
      text_chunks    : chunks RAG lisibles (tableaux Markdown + donnÃ©es)
       structured_data: données brutes pour Enno Valo (mapping Cerfa/templates)

    ParamÃ¨tres
    ----------
    file_path : str | Path
        Chemin vers le fichier .xlsx, .xlsm ou .xls

    Retourne
    --------
    ExcelStructResult
        text_chunks          : un chunk par feuille non-vide
        structured_data      : dict complet pour Valo
        sheets               : détail par feuille
        named_ranges         : plages nommées du classeur
        organisme_detected   : (reserve au NER)
        detected_rd_sections : sections CIR trouvées
        tags                 : traçabilité + catégories R&D
        confidence_score     : qualité globale

    Raises
    ------
    FileNotFoundError : fichier introuvable
    ValueError        : extension non supportée
    """
    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(f"Fichier introuvable : {path}")

    ext = path.suffix.lower()
    if ext not in (".xlsx", ".xlsm", ".xls"):
        raise ValueError(
            f"Extension '{ext}' non supportée. "
            f"Formats acceptés : .xlsx, .xlsm, .xls"
        )

    logger.info("Extraction Excel [%s] → %s", ext.upper(), path.name)

    if ext in (".xlsx", ".xlsm"):
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl non installÃ© : pip install openpyxl")
        result = _extract_with_openpyxl(path)
    else:
        # .xls ancien format
        if not XLRD_AVAILABLE:
            raise RuntimeError(
                "xlrd non installÃ© : pip install xlrd\n"
                "Note : xlrd >= 2.0 ne supporte que .xls (pas .xlsx)"
            )
        # Pour .xls on fait une conversion minimale via xlrd
        result = _extract_xls_fallback(path)

    result.tags = _build_tags(result)
    result.confidence_score = _compute_confidence(result)

    logger.info(
        "âœ“ %s â€” %d feuilles | %d chunks | %d tableaux | score=%.2f | tags=%s",
        path.name,
        len(result.sheets),
        len(result.text_chunks),
        sum(len(s.tables) for s in result.sheets),
        result.confidence_score,
        result.tags,
    )

    return result


def _extract_xls_fallback(path: Path) -> ExcelStructResult:
    """
    Extraction minimale pour les .xls anciens formats via xlrd.
    Produit uniquement des text_chunks (pas de structured_data riche).
    """
    result = ExcelStructResult(
        file_name=path.name,
        source_path=str(path.resolve()),
        file_type="excel_xls",
    )

    try:
        wb = xlrd.open_workbook(str(path))
    except Exception as exc:
        result.extraction_errors.append(f"Impossible d'ouvrir le .xls : {exc}")
        return result

    for sheet_idx in range(wb.nsheets):
        ws = wb.sheet_by_index(sheet_idx)
        sheet_name = ws.name
        rows_text: list[str] = []

        for row_idx in range(ws.nrows):
            row_vals = []
            for col_idx in range(ws.ncols):
                cell = ws.cell(row_idx, col_idx)
                val = cell.value
                if val not in (None, ""):
                    row_vals.append(str(val).strip())
            if len(row_vals) >= MIN_CELLS_FOR_DATA_ROW:
                rows_text.append(" | ".join(row_vals))

        if rows_text:
            chunk = f"[FEUILLE : {sheet_name}]\n\n" + "\n".join(rows_text)
            result.text_chunks.append(chunk)
            
            # Extraction des formules du chunk Excel
            formulas = extract_formulas(text=chunk)
            for formula in formulas:
                result.text_chunks.append(formula.rag_chunk)
            
            result.sheet_names.append(sheet_name)

    return result


# Interface rapide (debug / tests) 

if __name__ == "__main__":
    import sys
    import json

    logging.basicConfig(level=logging.DEBUG)

    if len(sys.argv) < 2:
        print("Usage : python excel_struct.py <chemin_vers_fichier.xlsx>")
        sys.exit(1)

    res = extract_excel(sys.argv[1])

    summary = {
        "file":          res.file_name,
        "sheets":        res.sheet_names,
        "chunks":        len(res.text_chunks),
        "organisme":     res.organisme_detected,
        "rd_sections":   res.detected_rd_sections,
        "named_ranges":  len(res.named_ranges),
        "tables_total":  sum(len(s.tables) for s in res.sheets),
        "rd_categories": list({
            cat for s in res.sheets for cat in s.rd_categories_found
        }),
        "tags":          res.tags,
        "confidence":    res.confidence_score,
        "errors":        res.extraction_errors,
        "chunks_preview": [c[:400] + "â€¦" for c in res.text_chunks[:2]],
        "structured_data_keys": list(res.structured_data.keys()),
    }

    print(json.dumps(summary, ensure_ascii=False, indent=2))

